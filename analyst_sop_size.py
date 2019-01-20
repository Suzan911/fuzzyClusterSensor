import os
import time
import openpyxl as xl
import multiprocessing as mp
import numpy as np
import config
from itertools import product
def get_avg_sop(size, t_init):
    SOP_avg = 0
    try:
        data = xl.load_workbook(config.root + "/R%02d/R%02dT%02ddata.xlsx" % (size, size, t_init))
    except:
        return

    arr = np.fromiter(map(lambda sheet: sheet.max_row - 1, data.worksheets), dtype=np.float64)
    SOP_avg = np.mean(arr, dtype=np.float64)
    size = []
    for sheet in data:
        size_per_rou = 0
        rou = 1
        for i in sheet["c":"c"]:#print(i.value) 
            if type(i.value) == type(1.2):
                size_per_rou += i.value
                rou += 1
        size += [size_per_rou/rou]
    a = 0
    tc = 1
    for i in size:
        a += i
        tc += 1
    size_avg = a/ tc
    return SOP_avg, size_avg
book = xl.Workbook()
sheet = book.active
sheet.title = "D%04d"%(config.root)
row = 1
sheet.cell(1, 1, 'size')
for r in config.size:
    row += 1
    col = 1
    sheet.cell(row, col, 'size%02i'%r)
    for t in config.t_init:
        SOP_avg, size_avg = get_avg_sop(r, t)
        sheet.cell(1, col*2, 'sop%02i'%t)
        sheet.cell(row, col*2+1, SOP_avg)
        sheet.cell(1, col*2, 'size%02i'%t)
        sheet.cell(row, col*2, size_avg)
        col +=1
book.save("%s.xlsx"%config.root)