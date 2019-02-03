"""
For running algorithm phase
"""
import os
import shutil
import openpyxl as xl
import time as time
import numpy as np
import multiprocessing as mp
import matplotlib.pyplot as plt
import matplotlib.markers as mark
import phase
import config
from writeExcel import writting
from itertools import product
from Field import Field
from Node import Node

def check_access_file(density, tc, size, t_init):
    """
    Check that have been already generate testcase or not
    """
    path = config.root + ("%.5f" % density) + ("_fuzzy" if is_fuzzy else "_fixed") + "/R%02d/T%d/%04d" % (density, size, t_init, tc)
    if not os.path.exists(path):
        return False, (tc, size, t_init)
    else:
        if os.path.exists(path + "/data.xlsx"):
            return True, (tc, size, t_init)
        else:
            shutil.rmtree(path)
            return False, (tc, size, t_init)


def running(density, tc, size, t_init, is_fuzzy=True):
    #----------------------
    # Initial value
    # Change these value if you want new property
    #----------------------
    t_init = t_init / 100
    path = config.root + ("%.5f" % density) + "_" + ("fuzzy" if is_fuzzy else "fixed") + ("/R%02d/T%02d/%04d" % (size, int(t_init * 100), tc))
    # size = 10
    #----------------------
    standy_loop = config.standy_loop
    field_radius = size

    start_time = time.time()
    print("Processing in {} testcase {} which set initial radius at {}, density at {} and T value at {}.\nRunning on processer {}\n".format(("fuzzy" if is_fuzzy else "fixed"), tc, field_radius, density, t_init, mp.current_process()))

    # Check if file already generate
    if not os.path.exists(path):
        os.makedirs(path)
    else:
        if os.path.exists(path + "/data.xlsx"):
            time_used = time.time() - start_time
            print("Processing at {} testcase {} which set initial radius at {}, density at {} and T value at {}.\nfinished within time {}s.\nRunning on processer {}\n".format(("fuzzy" if is_fuzzy else "fixed"), tc, field_radius, density, t_init, time_used, mp.current_process()))
            return
        else:
            try:
                shutil.rmtree(path)
            except Exception as err:
                pass
            os.makedirs(path)

    book = xl.Workbook()
    sheet = book.active
    sheet.title = "%04d" % tc
    sheet.cell(1, 1, "Round")
    sheet.cell(1, 2, "AVG_energy") 
    sheet.cell(1, 3, "Size")
    sheet.cell(1, 4, "T")
    sheet.cell(1, 5, "No Pointer node")

    field = Field(config.width, config.height, density, radius=field_radius, start_energy=3, t=t_init)
    left_node = [int(field.getDensity() * field.getWidth() * field.getHeight())]
    ignore_node = []
    while len(field.getNodes()) >= int(field.getDensity() * field.getWidth() * field.getHeight()):
        if phase.CCH_election_phase(field, t_init):
            # Running one setup phase
            phase.CH_competition_phase(field, field_radius)
            phase.cluster_announcement_phase(field, field_radius)
            phase.cluster_association_phase(field)
            phase.cluster_confirmation_phase(field, is_fuzzy=is_fuzzy, plot_graph=False)

            ignore_node = len(list(filter(lambda x: not x.hasPointerNode(), field.getNodes('CM'))))
            #field.printField(testcase=tc, showplot=0, rnd=field.getRound())

            # Data storage
            rnd = field.getRound()
            nodes = field.getNodes()
            CH_nodes = field.getNodes('CH')
            left_node.append(len(field.getNodes()))
            E_avg = (np.sum([n.getAverageAll_energy() for n in CH_nodes]) / len(CH_nodes)) if len(CH_nodes) else 0
            Size_avg = (np.sum([n.getSize() for n in CH_nodes]) / len(CH_nodes)) if len(CH_nodes) else 0
            T_avg = (np.sum([n.getT() for n in nodes]) / len(nodes)) if len(nodes) else 0

            sheet.cell(rnd + 1, 1, (len(left_node)-1))
            sheet.cell(rnd + 1, 2, E_avg)
            sheet.cell(rnd + 1, 3, Size_avg)
            sheet.cell(rnd + 1, 4, T_avg)
            sheet.cell(rnd + 1, 5, ignore_node)
            for _ in range(standy_loop):
                phase.standyPhase(field)
            field.nextRound()
            field.resetNode()

    e_avg_per_round = list(map(lambda cell: cell.value, sheet['B'][1:]))
    r_avg_per_round = list(map(lambda cell: cell.value, sheet['C'][1:]))
    t_avg_per_round = list(map(lambda cell: cell.value, sheet['D'][1:]))
    t_avg_case = np.mean(t_avg_per_round, dtype=np.float64)
    e_avg_case = np.mean(e_avg_per_round, dtype=np.float64)
    r_avg_case = np.mean(r_avg_per_round, dtype=np.float64)

    plt.plot(list(range(len(t_avg_per_round))), t_avg_per_round, linewidth=0.7, alpha=0.7)
    plt.plot([0, len(t_avg_per_round)], [t_avg_case, t_avg_case], color='red')
    plt.xlabel('Round')
    plt.ylabel('T')
    plt.title("T Average per round")
    # plt.show()
    plt.savefig(path + "/t_avg", dpi=300)
    plt.clf()

    plt.plot(list(range(len(e_avg_per_round))), e_avg_per_round, linewidth=0.7, alpha=0.7, color='green')
    plt.plot([0, len(e_avg_per_round)], [3, 0], color='red', linewidth=0.7, alpha=0.4)
    plt.xlabel('Round')
    plt.ylabel('Energy')
    plt.title("Energy Average per round")
    # plt.show()
    plt.savefig(path + "/energy_avg", dpi=300)
    plt.clf()

    plt.plot(list(range(len(r_avg_per_round))), r_avg_per_round, linewidth=0.7, alpha=0.7)
    plt.plot([0, len(r_avg_per_round)], [r_avg_case, r_avg_case], color='red')
    plt.xlabel('Round')
    plt.ylabel('Size Cluster')
    plt.title("Size Cluster Average per round")
    # plt.show()
    plt.savefig(path + "/size_avg", dpi=300)
    plt.clf()

    time_used = time.time() - start_time
    sheet.cell(1, 6, "Runtime (sec)")
    sheet.cell(2, 6, "%f" % time_used)

    try:
        book.save(path + "/data.xlsx")
        book.close()
    except Exception as err:
        print(err)

    print("Processing at {} testcase {} which set initial radius at {}, density at {} and T value at {}.\nfinished within time {}s.\nRunning on processer {}\n".format(("fuzzy" if is_fuzzy else "fixed"), tc, field_radius, density, t_init, time_used, mp.current_process()))


if __name__ == "__main__":
    """
    Initial Value
    """
    testcase = config.testcase
    t_initial = config.t_init
    size = config.size
    run_state = config.run_state

    """
    Check remaining testcase that not generate
    """
    validate_case = list(filter(lambda x: not x[0], [check_access_file(den, tc, s, t_value) for den in density for tc in testcase for t_value in t_initial for s in size]))
    chuck = []
    if run_state == "Fuzzy" or run_state == "Both":
        chuck.extend(sorted(list(map(lambda x, fuzzy=True: (*x[1], fuzzy), validate_case))))
    if run_state == "Fixed" or run_state == "Both":
        chuck.extend(sorted(list(map(lambda x, fuzzy=False: (*x[1], fuzzy), validate_case))))

    print("Currently there are", len(chuck), "testcase that not generate yet.\n")

    """
    Processing
    """
    start_time = time.time()
    pool = mp.Pool(config.pool)
    # Running thought T value for each 100 testcase
    pool.starmap(running, chuck) # product(testcase, t-initial, size)
    print("Finished simulate all testcase using {}s.".format(time.time() - start_time))

    """
    Writing the all of simulate into excel
    """
    writting()

    """
    Plot Graph
    To-do
    """
    #print("Starting plot graph...")
