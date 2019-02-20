"""
For running algorithm phase
"""
import os
import math
import shutil
import openpyxl as xl
import time as time
import numpy as np
import multiprocessing as mp
import matplotlib.pyplot as plt
import matplotlib.markers as mark
import phase
import config
from writeExcel import writting
from summarize import summarize
from itertools import product
from Field import Field
from Node import Node

def check_access_file(density, tc, size, t_init, is_fuzzy=True):
    """
    Check that have been already generate testcase or not
    """
    path = config.root + ("%.5f" % density) + ("_fuzzy" if is_fuzzy else "_fixed") + "/R%02d/T%02d/%04d" % (size, t_init, tc)
    if not os.path.exists(path + "/data.xlsx"):
        return False, (density, tc, size, t_init, is_fuzzy)
    else:
        return True, (density, tc, size, t_init, is_fuzzy)

def running(density, tc, size, t_init, is_fuzzy=True):
    #----------------------
    # Initial value
    # Change these value if you want new property
    #----------------------
    t_init = t_init / 100
    #----------------------
    standy_loop = config.standy_loop
    radius = size
    t_init_for_file = int(t_init * 100)
    field_radius = size
    path = config.root + ("%.5f" % density) + "_" + ("fuzzy" if is_fuzzy else "fixed") + ("/R%02d/T%02d/%04d" % (size, t_init_for_file, tc))
    

    start_time = time.time()
    print("Processing at {} testcase {} which set initial radius at {}, density at {} and T value at {}.\nRunning on processer {}\n".format(("fuzzy" if is_fuzzy else "fixed"), tc, field_radius, density, t_init, mp.current_process()))

    # Check if file already generate
    if not os.path.exists(path):
        os.makedirs(path)
    else:
        if os.path.exists(path + "/data.xlsx"):
            time_used = time.time() - start_time
            print("Processing at {} testcase {} which set initial radius at {}, density at {} and T value at {}.\nfinished within time {}s.\nRunning on processer {}\n".format(("fuzzy" if is_fuzzy else "fixed"), tc, field_radius, density, t_init, time_used, mp.current_process()))
            return
        else:
            try:
                shutil.rmtree(path)
            except Exception as err:
                pass
            os.makedirs(path)

    book = xl.Workbook()
    sheet = book.active
    sheet.title = "%04d" % tc
    sheet.cell(1, 1, "Round")
    sheet.cell(1, 2, "AVG_energy") 
    sheet.cell(1, 3, "Size")
    sheet.cell(1, 4, "HR")
    sheet.cell(1, 5, "No Pointer node")
    sheet.cell(1, 8, "Clustering")
    sheet.cell(1, 7, "Standy Phase")
    sheet.cell(1, 9, "Standy Phase AVG")
    sheet.cell(1, 10, "Clustering AVG")

    field = Field(config.width, config.height, density, radius=field_radius, start_energy=3, t=t_init)
    nodeLists = field.getNodes()
    left_node = [int(field.getDensity() * field.getWidth() * field.getHeight())]
    ignore_node = []
    alpha_radius = math.sqrt(2 * math.log(10)) * radius
    energy_memo = dict()
    Standy_AVG, ClusterAVG = 0, 0
    while field.nextRound():
        rnd = field.getRound()
        nodeList = field.getNodes()
        initEnergy = field.getInitEnergy()

        #print("\nRound", field.getRound())
        #----------------------------------------
        # Re-clustering ...
        #------------line2 T wait <- 1/Ej--------
        for node in nodeList:
            if node.getHyperRound() > 0:
                continue
            #plt.text(node.getX(), node.getY(), node.getEnergy(), fontsize=2, wrap=True)
            node.clearPointerNode()
            node.clearPackets()
            node.setType('CCH')
            node.setDelay(1/node.getEnergy()) # delay
            node.setState('active')
            energy_memo[node.getName()] = node.getEnergy()

        CCH_nodeList = sorted(field.getNodes('CCH'), key=lambda x: x.getDelay(), reverse=False)
        #print(CCH_nodeList)
        #------------line3 while T wait --------
        while len(CCH_nodeList):
            node = CCH_nodeList.pop(0)
            if not node.hasPointerNode():
                node.setType('CH')
                # Consume Energy for sended a packet
                node.consume_transmit(200, radius)
                for nearbyNode in field.getNearbyNodes(node, radius, 'CCH', debug=0):
                    # Consume Energy for received a packet
                    nearbyNode.consume_receive(200)
                    nearbyNode.setPointerNode(node)
            else:
                target = max(node.getPointerNode(), key=lambda n: n.getEnergy())
                node.consume_transmit(200, node.getDistanceFromNode(target))
                target.consume_receive(200)
                target.setPointerNode(node)
                node.setType('CM')

        #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-
        eper_r = 0
        for index, node in enumerate(field.getNodes(), start=1):
            before = energy_memo[node.getName()]
            after = node.getEnergy()
            eper_r += before - after
            energy_memo[node.getName()] = after
            ClusterAVG += eper_r
        sheet.cell(rnd + 1, 8, eper_r)
        #---------------Fuzzy Phase------------------
        CH_nodeList = field.getNodes('CH') # For each network node
        for node in CH_nodeList:
            if node.getHyperRound() <= 0:
                hyper_round = config.standy_loop
                if is_fuzzy:
                    network_node = [*node.getPointerNode(), node]
                    distance_node = list(map(lambda n, f=field: n.getDistanceFromNode(f.getBaseStation()), network_node))
                    if (node.hasPointerNode()):
                        relative_distance = min(max((node.getDistanceFromNode(field.getBaseStation()) - min(distance_node)) / (max(distance_node) - min(distance_node)), 0.1), 0.9)
                        relative_energy = min(max((node.getEnergy() / node.getInitEnergy()), 0.05), 0.95)
                        hyper_round = int(phase.HRFuzzy(relative_distance, relative_energy, 40))
                    #print("HR on", node.getName(), ":", hyper_round)

                node.setInitialHyperRound(hyper_round)
                for member in node.getPointerNode():
                    member.setInitialHyperRound(hyper_round)
        '''
        #-----print Field---------
        for node in field.getNodes('CH'):
            for member in node.getPointerNode():
                plt.plot([node.getX(), member.getX()], [node.getY(), member.getY()], color='r', alpha=0.4, linewidth=0.5)
        field.printField(path=path, testcase=tc, showplot=0, rnd=field.getRound())
        '''
        #---------------Standy Phase-----------------
        phase.standyPhase(field)
        #-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-#
        eper_r = 0
        rperr = 0
        for index, node in enumerate(field.getNodes(), start=1):
            before = energy_memo[node.getName()]
            after = node.getEnergy()
            eper_r += before - after
            Standy_AVG += eper_r
            energy_memo[node.getName()] = after
        sheet.cell(rnd + 1, 7, eper_r)
        #--------------------------------------------
        #ignore_node = len(list(filter(lambda x: not x.hasPointerNode(), field.getNodes('CM'))))
        
        # Data storage
        rnd = field.getRound()
        nodes = field.getNodes()
        CH_nodes = field.getNodes('CH')
        left_node.append(len(field.getNodes()))
        for node in CH_nodes:
            node.computeAverageEnergy()
            node.updateSize()

        E_avg = (np.sum([n.getAverageAll_energy() for n in CH_nodes]) / len(CH_nodes)) if len(CH_nodes) else 0
        Size_avg = (np.sum([n.getSize() for n in CH_nodes]) / len(CH_nodes)) if len(CH_nodes) else 0
        HR_avg = (np.sum([n.getInitialHyperRound() for n in CH_nodes]) / len(CH_nodes)) if len(CH_nodes) else 0

        sheet.cell(rnd + 1, 1, (len(left_node)-1))
        sheet.cell(rnd + 1, 2, E_avg)
        sheet.cell(rnd + 1, 3, Size_avg)
        sheet.cell(rnd + 1, 4, HR_avg)
    sheet.cell(2, 9, Standy_AVG/field.getRound()/len(nodeLists))
    sheet.cell(2, 10, ClusterAVG/field.getRound()/len(nodeLists))  

    e_avg_per_round = list(map(lambda cell: cell.value, sheet['B'][1:]))
    r_avg_per_round = list(map(lambda cell: cell.value, sheet['C'][1:]))
    hr_avg_per_round = list(map(lambda cell: cell.value, sheet['D'][1:]))
    es_avg_per_round = list(map(lambda cell: cell.value, sheet['G'][1:]))
    ec_avg_per_round = list(map(lambda cell: cell.value, sheet['H'][1:]))
    e_avg_case = np.mean(e_avg_per_round, dtype=np.float64)
    r_avg_case = np.mean(r_avg_per_round, dtype=np.float64)
    hr_avg_case = np.mean(hr_avg_per_round, dtype=np.float64)
    es_avg_case = np.mean(es_avg_per_round, dtype=np.float64)
    ec_avg_case = np.mean(ec_avg_per_round, dtype=np.float64)

    plt.plot(list(range(len(e_avg_per_round))), e_avg_per_round, linewidth=0.7, alpha=0.7, color='green')
    plt.plot([0, len(e_avg_per_round)], [3, 0], color='red', linewidth=0.7, alpha=0.4)
    plt.xlabel('Round')
    plt.ylabel('Energy')
    plt.title("Energy Average per round")
    # plt.show()
    plt.savefig(path + "/energy_avg", dpi=300)
    plt.clf()

    plt.plot(list(range(len(r_avg_per_round))), r_avg_per_round, linewidth=0.7, alpha=0.7)
    plt.plot([0, len(r_avg_per_round)], [r_avg_case, r_avg_case], color='red')
    plt.xlabel('Round')
    plt.ylabel('Size Cluster')
    plt.title("Size Cluster Average per round")
    # plt.show()
    plt.savefig(path + "/size_avg", dpi=300)
    plt.clf()

    plt.plot(list(range(len(hr_avg_per_round))), hr_avg_per_round, linewidth=0.7, alpha=0.7)
    plt.plot([0, len(hr_avg_per_round)], [hr_avg_case, hr_avg_case], color='red')
    plt.xlabel('Round')
    plt.ylabel('Hyper Round')
    plt.title("Hyper Round Average per round")
    # plt.show()
    plt.savefig(path + "/hr_avg", dpi=300)
    plt.clf()

    plt.plot(list(range(len(es_avg_per_round))), es_avg_per_round, linewidth=0.7, alpha=0.7)
    plt.plot([0, len(es_avg_per_round)], [es_avg_case, es_avg_case], color='red')
    plt.xlabel('Round')
    plt.ylabel('Energy Standy Phase Average per round')
    plt.title("Hyper Round Average per round")
    # plt.show()
    plt.savefig(path + "/es_avg", dpi=300)
    plt.clf()

    plt.plot(list(range(len(hr_avg_per_round))), hr_avg_per_round, linewidth=0.7, alpha=0.7)
    plt.plot([0, len(ec_avg_per_round)], [ec_avg_case, ec_avg_case], color='red')
    plt.xlabel('Round')
    plt.ylabel('Energy Clustering Average per round')
    plt.title("Hyper Round Average per round")
    # plt.show()
    plt.savefig(path + "/ec_avg", dpi=300)
    plt.clf()

    time_used = time.time() - start_time
    sheet.cell(1, 6, "Runtime (sec)")
    sheet.cell(2, 6, "%f" % time_used)

    try:
        book.save(path + "/data.xlsx")
        book.close()
    except Exception as err:
        print(err)

    print("Processing at {} testcase {} which set initial radius at {}, density at {} and T value at {}.\nfinished within time {}s.\nRunning on processer {}\n".format(("fuzzy" if is_fuzzy else "fixed"), tc, field_radius, density, t_init, time_used, mp.current_process()))


if __name__ == "__main__":
    """
    Initial Value
    """
    testcase = config.testcase
    t_initial = config.t_init
    size = config.size
    density = config.density
    state = []
    if config.run_state == "Fuzzy" or config.run_state == "Both":
        state.append(True)
    if config.run_state == "Fixed" or config.run_state == "Both":
        state.append(False)

    """
    Check remaining testcase that not generate
    """
    validate_case = list(filter(lambda x: not x[0], [check_access_file(den, tc, s, t_value, fuzz) 
                                                     for den in density for tc in testcase for t_value in t_initial for s in size for fuzz in state]))
    chuck = sorted(list(map(lambda x, fuzzy=True: x[1], validate_case)))
    print("Currently there are", len(chuck), "testcase that not generate yet.\n")

    """
    Processing
    """
    start_time = time.time()
    pool = mp.Pool(config.pool)
    # Running thought T value for each 100 testcase
    pool.starmap(running, chuck) # product(testcase, t-initial, size)
    print("Finished simulate all testcase using {}s.".format(time.time() - start_time))

    """
    Writing the all of simulate into excel
    """
    writting()

    """
    Plot Graph
    """
    summarize()
