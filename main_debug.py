"""
For running algorithm phase
"""
import os
import shutil
import openpyxl as xl
import time as time
import numpy as np
import multiprocessing as mp
import matplotlib.pyplot as plt
import matplotlib.markers as mark
import phase
import config
from writeExcel import writting
from itertools import product
from Field import Field
from Node import Node

def check_access_file(density, tc, size, t_init):
    """
    Check that have been already generate testcase or not
    """
    path = config.root + ("%.5f" % density) + ("_fuzzy" if is_fuzzy else "_fixed") + "/R%02d/T%d/%04d" % (density, size, t_init, tc)
    if not os.path.exists(path):
        return False, (tc, size, t_init)
    else:
        if os.path.exists(path + "/data.xlsx"):
            return True, (tc, size, t_init)
        else:
            shutil.rmtree(path)
            return False, (tc, size, t_init)


# This is main
def running(density, tc, size, t_init, is_fuzzy=True):
    #----------------------
    # Initial value
    # Change these value if you want new property
    #----------------------
    t_init = t_init / 100
    path = config.root + ("%.5f" % density) + "_" + ("fuzzy" if is_fuzzy else "fixed") + ("/R%02d/T%02d/%04d" % (size, int(t_init * 100), tc))
    # size = 10
    #----------------------
    standy_loop = config.standy_loop
    field_radius = size

    start_time = time.time()
    print("Processing in {} testcase {} which set initial radius at {}, density at {} and T value at {}.\nRunning on processer {}\n".format(("fuzzy" if is_fuzzy else "fixed"), tc, field_radius, density, t_init, mp.current_process()))

    # Check if file already generate
    if not os.path.exists(path):
        os.makedirs(path)
    else:
        if os.path.exists(path + "/data.xlsx"):
            time_used = time.time() - start_time
            print("Processing at {} testcase {} which set initial radius at {}, density at {} and T value at {}.\nfinished within time {}s.\nRunning on processer {}\n".format(("fuzzy" if is_fuzzy else "fixed"), tc, field_radius, density, t_init, time_used, mp.current_process()))
            return
        else:
            try:
                shutil.rmtree(path)
            except Exception as err:
                pass
            os.makedirs(path)

    book = xl.Workbook()
    sheet = book.active
    sheet.title = "%04d" % tc
    sheet.cell(1, 1, "Round")
    sheet.cell(1, 2, "AVG_energy") 
    sheet.cell(1, 3, "Size")
    sheet.cell(1, 4, "T")
    sheet.cell(1, 5, "No Pointer node")

    sheet_energy = book.create_sheet("energy")
    sheet_energy.cell(1, 1, "Round")
    sheet_energy.cell(1, 2, "Phase")

    field = Field(100, 100, density, radius=field_radius, start_energy=3, t=t_init)
    left_node = [int(field.getDensity() * field.getWidth() * field.getHeight())]
    t_avg_per_round = []
    e_avg_per_round = []
    r_avg_per_round = []
    ignore_node = []

    energy_memo = dict()
    for index, node in enumerate(field.getNodes(), start=1):
        sheet_energy.cell(1, index + 3, "Node (%.2f, %.2f)"% node.getPosition())
        energy_memo[node.getName()] = node.getEnergy()

    while len(field.getNodes()) >= (field.getDensity() * field.getWidth() * field.getHeight()):
        rnd = field.getRound()
        print(rnd)
        _step = 12
        # Phase 1 :: Loop until have at least one CCH
        if phase.CCH_election_phase(field, t_init):
            # Running one setup phase
            sheet_energy.cell((rnd - 1) * _step + 2, 1, rnd)
            sheet_energy.cell((rnd - 1) * _step + 2, 2, "Current Energy")
            for index, node in enumerate(field.getNodes(), start=1):
                sheet_energy.cell((rnd - 1) * _step + 2, index + 3, energy_memo[node.getName()])

            # Phase 2
            phase2_debug = phase.CH_competition_phase(field, field_radius, debug=True)
            sheet_energy.cell((rnd - 1) * _step + 3, 2, "CH_competition_phase")
            for index, node in enumerate(field.getNodes(), start=1):
                before = energy_memo[node.getName()]
                after = node.getEnergy()
                sheet_energy.cell((rnd - 1) * _step + 3, index + 3, before - after)
                energy_memo[node.getName()] = after

            # Phase 3
            phase3_debug = phase.cluster_announcement_phase(field, field_radius, debug=True)
            sheet_energy.cell((rnd - 1) * _step + 4, 2, "cluster_announcement_phase")
            for index, node in enumerate(field.getNodes(), start=1):
                before = energy_memo[node.getName()]
                after = node.getEnergy()
                sheet_energy.cell((rnd - 1) * _step + 4,  index + 3, before - after)
                energy_memo[node.getName()] = after

            # Phase 4
            phase4_debug = phase.cluster_association_phase(field, debug=True)
            sheet_energy.cell((rnd - 1) * _step + 5, 2, "cluster_association_phase")
            for index, node in enumerate(field.getNodes(), start=1):
                before = energy_memo[node.getName()]
                after = node.getEnergy()
                sheet_energy.cell((rnd - 1) * _step + 5, index + 3, before - after)
                energy_memo[node.getName()] = after

            # Phase 5
            phase.cluster_confirmation_phase(field, is_fuzzy=is_fuzzy, plot_graph=False)
            sheet_energy.cell((rnd - 1) * _step + 6, 2, "cluster_confirmation_phase")
            for index, node in enumerate(field.getNodes(), start=1):
                before = energy_memo[node.getName()]
                after = node.getEnergy()
                sheet_energy.cell((rnd - 1) * _step + 6, index + 3, before - after)
                energy_memo[node.getName()] = after
                ignore_node = len(list(filter(lambda x: not x.hasPointerNode(), field.getNodes('CM'))))

            ignore_node = len(list(filter(lambda x: not x.hasPointerNode(), field.getNodes('CM'))))
            #field.printField(testcase=tc, showplot=0, rnd=field.getRound())
            
            # Data storage
            nodes = field.getNodes()
            CH_nodes = field.getNodes('CH')
            left_node.append(len(field.getNodes()))
            E_avg = (np.sum([n.getAverageAll_energy() for n in CH_nodes]) / len(CH_nodes)) if len(CH_nodes) else 0
            Size_avg = (np.sum([n.getSize() for n in CH_nodes]) / len(CH_nodes)) if len(CH_nodes) else 0
            T_avg = (np.sum([n.getT() for n in nodes]) / len(nodes)) if len(nodes) else 0

            e_avg_per_round.append(E_avg)
            r_avg_per_round.append(Size_avg)
            t_avg_per_round.append(T_avg)
            sheet.cell(rnd + 1, 1, (len(left_node)-1))
            sheet.cell(rnd + 1, 2, E_avg)
            sheet.cell(rnd + 1, 3, Size_avg)
            sheet.cell(rnd + 1, 4, T_avg)
            sheet.cell(rnd + 1, 5, ignore_node)

            for r in range(1, standy_loop + 1):
                phase.standyPhase(field)


            field.nextRound()
            field.resetNode()
    
    # Save graph
    '''
    e_avg_per_round = list(map(lambda cell: cell.value, sheet['B'][1:]))
    r_avg_per_round = list(map(lambda cell: cell.value, sheet['C'][1:]))
    t_avg_per_round = list(map(lambda cell: cell.value, sheet['D'][1:]))
    t_avg_case = np.mean(t_avg_per_round, dtype=np.float64)
    e_avg_case = np.mean(e_avg_per_round, dtype=np.float64)
    r_avg_case = np.mean(r_avg_per_round, dtype=np.float64)

    plt.plot(list(range(len(t_avg_per_round))), t_avg_per_round, linewidth=0.7, alpha=0.7)
    plt.plot([0, len(t_avg_per_round)], [t_avg_case, t_avg_case], color='red')
    plt.xlabel('Round')
    plt.ylabel('T')
    plt.title("T Average per round")
    # plt.show()
    plt.savefig(path + "/t_avg", dpi=300)
    plt.clf()

    plt.plot(list(range(len(e_avg_per_round))), e_avg_per_round, linewidth=0.7, alpha=0.7, color='green')
    plt.plot([0, len(e_avg_per_round)], [3, 0], color='red', linewidth=0.7, alpha=0.4)
    plt.xlabel('Round')
    plt.ylabel('Energy')
    plt.title("Energy Average per round")
    # plt.show()
    plt.savefig(path + "/energy_avg", dpi=300)
    plt.clf()

    plt.plot(list(range(len(r_avg_per_round))), r_avg_per_round, linewidth=0.7, alpha=0.7)
    plt.plot([0, len(r_avg_per_round)], [r_avg_case, r_avg_case], color='red')
    plt.xlabel('Round')
    plt.ylabel('Size Cluster')
    plt.title("Size Cluster Average per round")
    # plt.show()
    plt.savefig(path + "/size_avg", dpi=300)
    plt.clf()   
    '''

    time_used = time.time() - start_time
    sheet.cell(1, 6, "Runtime (sec)")
    sheet.cell(2, 6, "%f" % time_used)

    try:
        book.save(path + "/data.xlsx")
        book.close()
    except Exception as err:
        print(err)
    del book

    print("Processing at testcase {} which set initial radius at {}, density at {} and T value at {}.\nfinished within time {}s.\nRunning on processer {}\n".format(tc, field_radius, density, t_init, time_used, mp.current_process()))


if __name__ == "__main__":
    """
    Initial Value
    """
    testcase = config.testcase
    t_initial = config.t_init
    size = config.size
    density = config.density
    run_state = config.run_state
    print("=== Running on debugging mode! ===")

    """
    Check remaining testcase that not generate
    """
    validate_case = list(filter(lambda x: not x[0], [check_access_file(den, tc, s, t_value) for den in density for tc in testcase for t_value in t_initial for s in size]))
    chuck = []
    if run_state == "Fuzzy" or run_state == "Both":
        chuck.extend(sorted(list(map(lambda x, fuzzy=True: (*x[1], fuzzy), validate_case))))
    if run_state == "Fixed" or run_state == "Both":
        chuck.extend(sorted(list(map(lambda x, fuzzy=False: (*x[1], fuzzy), validate_case))))
    print("Currently there are", len(chuck), "testcase that not generate yet.\n")

    """
    Processing
    """        
    start_time = time.time()
    pool = mp.Pool(config.pool)
    # Running thought T value for each 100 testcase
    pool.starmap(running, chuck) # product(testcase, t-initial, size)
    print("Finished simulate all testcase using {}s.".format(time.time() - start_time))
    
    """
    Writing the all of simulate into excel
    """
    writting()
