import os
import time
import openpyxl as xl
import multiprocessing as mp
import numpy as np
import config
from itertools import product

def readExcelFile(tc, t_init, size, output=0):
    # If file not exist
    data = validate_access_file(tc, t_init, size)
    if not data:
        return 0
    data = data[1]

    start_time = time.time()
    try:
        book = xl.load_workbook(config.root + "/R%02d/R%02dT%02ddata.xlsx" % (size, size, t_init))
        sheet = book.create_sheet("%04d" % tc)
    except:
        book = xl.Workbook()
        sheet = book.active
        sheet.title = "%04d" % tc

    table = data["%04d" % tc].rows
    for row in table:
        for cell in row:
            sheet[cell.coordinate] = cell.value

    book.save(config.root + "/R%02d/R%02dT%02ddata.xlsx" % (size, size, t_init))
    book.close()

    if output:
        print("Processing at testcase {} which set initial radius at {} and T value at {}\nfinished within time {} s.\nRunning on processer {}\n".format(tc, size, t_init_for_file / 100, time.time() - start_time, mp.current_process()))
    return 1


def validate_access_file(tc, t_init, size):
    try:
        data = xl.load_workbook(config.root + "/R%02d/T%02d/%04d/data.xlsx" % (size, t_init, tc))
        return (tc, t_init, size), data
    except:
        return 0


if __name__ == "__main__":
    """
    Initial Value
    """
    testcase = config.testcase
    t_initial = config.t_init
    size = config.size
    output = True

    """
    Running
    """
    pool = mp.Pool(config.pool)
    # Running thought T value for each 100 testcase
    pool.starmap(readExcelFile, product(testcase, t_initial, size, [output])) # product(testcase, t-initial, size)
