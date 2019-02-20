import os
import time
import openpyxl as xl
import multiprocessing as mp
import numpy as np
import config
from itertools import product

""" Duplicate """
def readExcelFile(density, tc, t_init, size, root_path=config.root, output=0):
    # If file not exist
    root_path = config.root + ("%.5f" % density) + ("_fuzzy" if is_fuzzy else "_fixed") + ("/R%02d" % (size))
    data = validate_access_file(tc, t_init, size, path)
    if not data:
        return 1
    data = data[2]

    start_time = time.time()
    try:
        book = xl.load_workbook(root_path + "/R%02d/R%02dT%02ddata.xlsx" % (size, size, t_init))
        sheet = book.create_sheet("%04d" % tc)
    except:
        book = xl.Workbook()
        sheet = book.active
        sheet.title = "%04d" % tc

    table = data["%04d" % tc].rows
    for row in table:
        for cell in row:
            sheet[cell.coordinate] = cell.value

    book.save(root_path + "/R%02d/R%02dT%02ddata.xlsx" % (size, size, t_init))
    book.close()

    if output:
        print("Processing at testcase {} which set initial radius at {} and T value at {}\nfinished within time {} s.\nRunning on processer {}\n".format(tc, size, t_init_for_file / 100, time.time() - start_time, mp.current_process()))
    return 1


def writeExcelFile(density, t_init, size, is_fuzzy=True):
    start_time = time.time()
    root_path = config.root + ("%.5f" % density) + ("_fuzzy" if is_fuzzy else "_fixed") + ("/R%02d" % (size))
    try:
        book = xl.load_workbook(root_path + "/R%02dT%02ddata.xlsx" % (size, t_init))
    except:
        book = xl.Workbook()
        book.remove_sheet(book.active)

    for tc in config.testcase:
        if tc not in book.worksheets:
            path = root_path + "/T%02d/%04d/data.xlsx" % (t_init, tc)
            data = validate_access_file(density, tc, t_init, size, path)
            if data:
                data = data[2]
                sheet = book.create_sheet("%04d" % tc)
                table = data["%04d" % tc].rows
                for row in table:
                    for cell in row:
                        sheet[cell.coordinate] = cell.value

    book.save(root_path + "/R%02dT%02ddata.xlsx" % (size, t_init))
    book.close()
    print("Process simulation which set initial radius at {} and T value at {}\nfinished within time {} s.\nRunning on processer {}\n".format(size, t_init / 100, time.time() - start_time, mp.current_process()))
    return 1


def check_access_file(density, tc, size, t_init, state):
    """
    Check that have been already generate testcase or not
    """
    path = config.root + ("%.5f" % density) + ("_fuzzy" if state else "_fixed") + ("/R%02d/T%02d/%04d" % (size, t_init, tc))
    if not os.path.exists(path):
        return False, (tc, size, t_init)
    else:
        if os.path.exists(path + "/data.xlsx"):
            return True, (tc, size, t_init)
        else:
            shutil.rmtree(path)
            return False, (tc, size, t_init)


def validate_access_file(density, tc, t_init, size, data_path):
    try:
        data = xl.load_workbook(data_path)
        return True, (tc, t_init, size, data_path), data
    except Exception as err:
        print(err)
        return False


def writting():
    """
    Initial Value
    """
    testcase = config.testcase
    density = config.density
    t_initial = config.t_init
    size = config.size
    run_state = []
    if config.run_state == "Fuzzy" or config.run_state == "Both":
        run_state.append(True)
    if config.run_state == "Fixed" or config.run_state == "Both":
        run_state.append(False)

    """
    Check remaining testcase that not generate
    """
    chuck = [(den, t_init, siz, state) for den in density for t_init in t_initial for siz in size for state in run_state]

    """
    Running
    """
    pool = mp.Pool(config.pool)
    # Running thought T value for each 100 testcase
    pool.starmap(writeExcelFile, chuck) # product(testcase, t-initial, size)


if __name__ == "__main__":
    writting()
